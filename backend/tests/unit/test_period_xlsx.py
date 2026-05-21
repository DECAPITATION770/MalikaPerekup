"""Period report → .xlsx serialization.

Pure (no DB): build a workbook from a known PeriodReport, load it back
with openpyxl and assert the cells the user/accountant will read. The
regression-worthy part is that numbers land as numbers (not strings) and
localisation picks the right labels (CLAUDE.md §13).
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.features.reports.export import build_period_xlsx
from app.features.reports.schemas import DayProfit, PeriodReport, TopModelEntry

REPORT = PeriodReport(
    date_from=date(2026, 4, 19),
    date_to=date(2026, 5, 19),
    purchases_count=8,
    purchases_total_uzs=Decimal("73500000.00"),
    sales_count=12,
    revenue_uzs=Decimal("92000000.00"),
    profit_uzs=Decimal("18500000.00"),
    avg_profit_per_sale_uzs=Decimal("1541667.00"),
    returns_count=1,
    top_models=[
        TopModelEntry(
            brand="Apple", model="iPhone 14 Pro",
            units_sold=5, total_profit_uzs=Decimal("7500000.00"),
        ),
    ],
    avg_days_in_stock=8.34,
    profit_by_day=[],
)


def _load(report=REPORT, lang="ru"):
    return load_workbook(BytesIO(build_period_xlsx(report, lang=lang))).active


def test_xlsx_is_valid_and_carries_period():
    ws = _load()
    assert ws["A1"].value == "Отчёт"
    assert ws["B2"].value == "2026-04-19 — 2026-05-19"


def test_money_cells_are_numeric_not_strings():
    ws = _load()
    # Find the profit row by its label, assert the adjacent value is a
    # real number so Excel can sum/sort it.
    profit_val = None
    for row in ws.iter_rows(min_col=1, max_col=2):
        if row[0].value == "Прибыль (UZS)":
            profit_val = row[1].value
            break
    assert profit_val == 18500000
    assert isinstance(profit_val, (int, float))


def test_top_models_block_present():
    ws = _load()
    flat = [v for col in ws.iter_cols(values_only=True) for v in col]
    assert "Apple iPhone 14 Pro" in flat
    assert 5 in flat
    assert 7500000 in flat


def test_uzbek_localisation():
    ws = _load(lang="uz")
    assert ws["A1"].value == "Hisobot"
    flat = [v for col in ws.iter_cols(values_only=True) for v in col]
    assert "Foyda (UZS)" in flat
    assert "Top modellar" in flat


def test_avg_days_none_renders_dash():
    r = REPORT.model_copy(update={"avg_days_in_stock": None, "top_models": []})
    ws = _load(r)
    flat = [v for col in ws.iter_cols(values_only=True) for v in col]
    assert "—" in flat


def test_daily_series_lands_on_second_sheet():
    """Non-empty profit_by_day → second sheet "По дням" with day+profit rows."""
    r = REPORT.model_copy(update={"profit_by_day": [
        DayProfit(day=date(2026, 5, 17), profit_uzs=Decimal("300000.00")),
        DayProfit(day=date(2026, 5, 18), profit_uzs=Decimal("0.00")),
        DayProfit(day=date(2026, 5, 19), profit_uzs=Decimal("450000.00")),
    ]})
    wb = load_workbook(BytesIO(build_period_xlsx(r, lang="ru")))
    assert "По дням" in wb.sheetnames
    ws2 = wb["По дням"]
    # Header row + 3 days. openpyxl reads dates back as datetime; compare
    # the parts to stay agnostic about that.
    assert ws2.cell(row=1, column=1).value == "День"
    d1 = ws2.cell(row=2, column=1).value
    assert (d1.year, d1.month, d1.day) == (2026, 5, 17)
    assert ws2.cell(row=2, column=2).value == 300000
    assert ws2.cell(row=4, column=2).value == 450000


def test_empty_daily_series_skips_second_sheet():
    wb = load_workbook(BytesIO(build_period_xlsx(REPORT, lang="ru")))
    assert wb.sheetnames == ["Отчёт"]
