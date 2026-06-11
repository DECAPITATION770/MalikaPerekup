"""Tabular Excel export with per-column selection. Tested against real
PostgreSQL: column subset + order are honoured, rows are shop-scoped (§6),
and the workbook opens with the expected header row."""

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.features.admin import service as admin_service
from app.features.purchases import service as purchase_service
from app.features.reports import export_table
from app.features.sales import service as sale_service


class _DeviceIn:
    def __init__(self, imei, brand="Apple", model="iPhone 13"):
        self.category = "phone"; self.brand = brand; self.model = model
        self.imei = imei; self.serial = None; self.condition = "good"
        self.specs = {}; self.photos = []; self.defects = []; self.notes = None


class _PersonIn:
    def __init__(self, full_name="P"):
        self.full_name = full_name; self.phone = None; self.doc_type = None
        self.doc_number = None; self.photos = []; self.tg_username = None


async def _shop(db, name, tg):
    shop, user = await admin_service.register_shop_with_owner(
        db, name=name, language_default="ru", owner_full_name=f"O{name}",
        owner_tg_id=tg, owner_tg_username=None, owner_phone=None,
        owner_login=None, owner_password=None)
    await db.flush()
    return shop, user


async def _sell(db, shop, user, *, imei, buy, sell, day, brand="Apple"):
    _, dev = await purchase_service.create_purchase(
        db, shop_id=shop.id, user_id=user.id, device_in=_DeviceIn(imei, brand=brand),
        seller_in=_PersonIn(), currency="UZS", price=Decimal(buy),
        exchange_rate=None, purchase_date=day - timedelta(days=2), comment=None)
    await sale_service.create_sale(
        db, shop_id=shop.id, user_id=user.id, device_id=dev.id,
        buyer_in=_PersonIn("Buyer"), sale_type="cash", currency="UZS",
        price=Decimal(sell), exchange_rate=None, sale_date=day, comment=None)


def _open(xlsx: bytes):
    return load_workbook(BytesIO(xlsx)).active


async def test_sales_export_respects_selected_columns_and_order(db):
    shop, user = await _shop(db, "A", 9701)
    day = date(2026, 5, 19)
    await _sell(db, shop, user, imei="s1", buy="1000000", sell="1300000", day=day)

    rows = await export_table.fetch(db, shop_id=shop.id, entity="sales",
                                    date_from=day, date_to=day)
    # caller picks only these three, in this order
    xlsx = export_table.build_xlsx("sales", rows, ["date", "profit", "brand"], lang="ru")
    ws = _open(xlsx)

    assert [c.value for c in ws[1]] == ["Дата", "Прибыль (UZS)", "Бренд"]
    # openpyxl reads a date cell back as datetime — compare the date part.
    assert ws[2][0].value.date() == day
    assert ws[2][1].value == 300000.0   # profit, money → float
    assert ws[2][2].value == "Apple"


async def test_unknown_column_keys_are_ignored(db):
    shop, user = await _shop(db, "A", 9702)
    day = date(2026, 5, 19)
    await _sell(db, shop, user, imei="s1", buy="500000", sell="700000", day=day)
    rows = await export_table.fetch(db, shop_id=shop.id, entity="sales",
                                    date_from=day, date_to=day)
    xlsx = export_table.build_xlsx("sales", rows, ["brand", "does_not_exist"], lang="ru")
    ws = _open(xlsx)
    assert [c.value for c in ws[1]] == ["Бренд"]  # bogus key dropped


async def test_devices_export_is_shop_isolated(db):
    shop_a, ua = await _shop(db, "A", 9703)
    shop_b, ub = await _shop(db, "B", 9704)
    day = date(2026, 5, 19)
    await _sell(db, shop_a, ua, imei="a1", buy="1000000", sell="1300000", day=day, brand="Apple")
    await _sell(db, shop_b, ub, imei="b1", buy="900000", sell="1000000", day=day, brand="Samsung")

    rows = await export_table.fetch(db, shop_id=shop_a.id, entity="devices",
                                    date_from=None, date_to=None)
    xlsx = export_table.build_xlsx("devices", rows, ["brand", "imei"], lang="ru")
    ws = _open(xlsx)
    brands = [r[0].value for r in ws.iter_rows(min_row=2)]
    assert brands == ["Apple"]  # shop B's Samsung never appears


async def test_columns_metadata_lists_keys_and_labels():
    cols = export_table.columns_for("purchases", lang="ru")
    keys = {c["key"] for c in cols}
    assert {"date", "brand", "seller", "price_uzs"} <= keys
    assert all("label" in c and "type" in c for c in cols)
