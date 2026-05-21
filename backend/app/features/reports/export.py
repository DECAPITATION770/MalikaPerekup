"""Render a :class:`PeriodReport` into a real ``.xlsx`` workbook.

Mirrors the qr.png pattern: the backend owns binary-artifact generation
so the data stays shop-scoped at the source and the Mini App bundle
stays lean (no SheetJS). One sheet: a summary block + a top-models block.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.features.reports.schemas import PeriodReport

_MONEY_FMT = "#,##0"

_LABELS: dict[str, dict[str, str]] = {
    "ru": {
        "title": "Отчёт",
        "period": "Период",
        "metric": "Показатель",
        "value": "Значение",
        "purchases": "Закупок",
        "purchases_sum": "Сумма закупок (UZS)",
        "sales": "Продаж",
        "revenue": "Выручка (UZS)",
        "profit": "Прибыль (UZS)",
        "avg_profit": "Средняя прибыль/продажа (UZS)",
        "returns": "Возвратов",
        "avg_days": "Ср. дней на витрине",
        "top_models": "Топ моделей",
        "model": "Модель",
        "units": "Продано",
        "daily_sheet": "По дням",
        "day": "День",
    },
    "uz": {
        "title": "Hisobot",
        "period": "Davr",
        "metric": "Ko'rsatkich",
        "value": "Qiymat",
        "purchases": "Qabul",
        "purchases_sum": "Qabul summasi (UZS)",
        "sales": "Sotuvlar",
        "revenue": "Tushum (UZS)",
        "profit": "Foyda (UZS)",
        "avg_profit": "O'rtacha foyda/sotuv (UZS)",
        "returns": "Qaytarishlar",
        "avg_days": "O'rtacha vitrinada (kun)",
        "top_models": "Top modellar",
        "model": "Model",
        "units": "Sotildi",
        "daily_sheet": "Kunlar bo'yicha",
        "day": "Kun",
    },
}


def _money(value: Decimal | str | None) -> float:
    if value is None or value == "":
        return 0.0
    return float(Decimal(str(value)))


def build_period_xlsx(report: PeriodReport, *, lang: str = "ru") -> bytes:
    """Return the period report as ``.xlsx`` bytes ready to stream/download."""
    L = _LABELS.get(lang, _LABELS["ru"])

    wb = Workbook()
    ws = wb.active
    ws.title = L["title"]

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws["A1"] = L["title"]
    ws["A1"].font = title_font
    ws["A2"] = L["period"]
    ws["A2"].font = bold
    ws["B2"] = f"{report.date_from.isoformat()} — {report.date_to.isoformat()}"

    # ── Summary table ──
    row = 4
    ws.cell(row=row, column=1, value=L["metric"]).font = bold
    ws.cell(row=row, column=2, value=L["value"]).font = bold
    row += 1

    summary: list[tuple[str, object, bool]] = [
        (L["purchases"], report.purchases_count, False),
        (L["purchases_sum"], _money(report.purchases_total_uzs), True),
        (L["sales"], report.sales_count, False),
        (L["revenue"], _money(report.revenue_uzs), True),
        (L["profit"], _money(report.profit_uzs), True),
        (L["avg_profit"], _money(report.avg_profit_per_sale_uzs), True),
        (L["returns"], report.returns_count, False),
        (
            L["avg_days"],
            round(report.avg_days_in_stock, 1)
            if report.avg_days_in_stock is not None
            else "—",
            False,
        ),
    ]
    for label, value, is_money in summary:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if is_money:
            cell.number_format = _MONEY_FMT
        row += 1

    # ── Top models ──
    if report.top_models:
        row += 1
        ws.cell(row=row, column=1, value=L["top_models"]).font = bold
        row += 1
        for col, head in enumerate(
            (L["model"], L["units"], L["profit"]), start=1
        ):
            ws.cell(row=row, column=col, value=head).font = bold
        row += 1
        for m in report.top_models:
            ws.cell(row=row, column=1, value=f"{m.brand} {m.model}")
            ws.cell(row=row, column=2, value=m.units_sold)
            pc = ws.cell(row=row, column=3, value=_money(m.total_profit_uzs))
            pc.number_format = _MONEY_FMT
            row += 1

    # Readable column widths.
    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 18
    ws["A1"].alignment = Alignment(vertical="center")

    # ── Second sheet: raw daily series for accountant-side pivoting ──
    if report.profit_by_day:
        ws2 = wb.create_sheet(title=L["daily_sheet"])
        ws2.cell(row=1, column=1, value=L["day"]).font = bold
        ws2.cell(row=1, column=2, value=f"{L['profit']} (UZS)").font = bold
        for i, point in enumerate(report.profit_by_day, start=2):
            ws2.cell(row=i, column=1, value=point.day)
            ws2.cell(row=i, column=1).number_format = "yyyy-mm-dd"
            pc = ws2.cell(row=i, column=2, value=_money(point.profit_uzs))
            pc.number_format = _MONEY_FMT
        ws2.column_dimensions[get_column_letter(1)].width = 14
        ws2.column_dimensions[get_column_letter(2)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
