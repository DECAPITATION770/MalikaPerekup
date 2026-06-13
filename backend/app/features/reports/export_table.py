"""Plain tabular Excel export with per-column selection.

Unlike ``export.py`` (a styled summary report), this produces a flat sheet —
one header row, one row per record, only the columns the user ticked, in the
order they asked for. Powers Reports → "Выгрузка в Excel".

Each entity (sales / devices / purchases) declares an ordered list of columns;
a column knows its bilingual header, its cell type (for number formats), and
how to pull its value out of a result row. The frontend reads the same
registry via :func:`columns_for` so the checkbox UI never drifts from what the
export can actually produce.

Everything is shop-scoped at the query (CLAUDE.md §6); money stays Decimal
until the final float cast openpyxl needs.
"""

from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import Row, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.dates import to_tashkent
from app.features.devices.models import Device, DeviceStatus
from app.features.purchases.models import Purchase
from app.features.sales.models import Sale, SaleStatus

_MONEY_FMT = "#,##0"
_DATE_FMT = "yyyy-mm-dd"
_DATETIME_FMT = "yyyy-mm-dd hh:mm"

# ── Localised enum value maps (raw value → display). Fallback = raw value. ──
_ENUM: dict[str, dict[str, dict[str, str]]] = {
    "category": {
        "ru": {"phone": "Телефон", "tablet": "Планшет", "laptop": "Ноутбук",
               "smartwatch": "Часы", "accessory": "Аксессуар", "other": "Другое"},
        "uz": {"phone": "Telefon", "tablet": "Planshet", "laptop": "Noutbuk",
               "smartwatch": "Soat", "accessory": "Aksessuar", "other": "Boshqa"},
    },
    "condition": {
        "ru": {"new": "Новый", "good": "Хорошее", "normal": "Нормальное", "broken": "Битый"},
        "uz": {"new": "Yangi", "good": "Yaxshi", "normal": "O‘rtacha", "broken": "Buzuq"},
    },
    "device_status": {
        "ru": {"in_stock": "На витрине", "reserved": "Бронь", "sold": "Продан",
               "returned": "Возврат", "written_off": "Списан"},
        "uz": {"in_stock": "Vitrinada", "reserved": "Bron", "sold": "Sotilgan",
               "returned": "Qaytarilgan", "written_off": "Hisobdan chiqarilgan"},
    },
    "sale_status": {
        "ru": {"active": "Активна", "returned": "Возврат", "cancelled": "Отменена"},
        "uz": {"active": "Faol", "returned": "Qaytarilgan", "cancelled": "Bekor qilingan"},
    },
    "sale_type": {
        "ru": {"cash": "Наличные", "nasiya": "Рассрочка"},
        "uz": {"cash": "Naqd", "nasiya": "Nasiya"},
    },
}


def _enum(group: str, value: str | None, lang: str) -> str:
    if value is None:
        return ""
    return _ENUM.get(group, {}).get(lang, {}).get(value, value)


@dataclass(frozen=True)
class Col:
    key: str
    ru: str
    uz: str
    type: str  # text | int | money | date | datetime
    get: Callable[[Row, str], Any]

    def header(self, lang: str) -> str:
        return self.uz if lang == "uz" else self.ru


# ── Column registries ──────────────────────────────────────────────────
# Order here is the default column order in the UI.

_SALES_COLS: list[Col] = [
    Col("date", "Дата", "Sana", "date", lambda r, l: r.sale_date),
    Col("brand", "Бренд", "Brend", "text", lambda r, l: r.brand),
    Col("model", "Модель", "Model", "text", lambda r, l: r.model),
    Col("imei", "IMEI", "IMEI", "text", lambda r, l: r.imei or ""),
    Col("category", "Категория", "Toifa", "text", lambda r, l: _enum("category", r.category, l)),
    Col("condition", "Состояние", "Holati", "text", lambda r, l: _enum("condition", r.condition, l)),
    Col("buyer", "Покупатель", "Xaridor", "text", lambda r, l: r.buyer_name),
    Col("buyer_phone", "Телефон", "Telefon", "text", lambda r, l: r.buyer_phone or ""),
    Col("payment", "Оплата", "To‘lov", "text", lambda r, l: _enum("sale_type", r.sale_type, l)),
    Col("price_uzs", "Цена продажи (UZS)", "Sotuv narxi (UZS)", "money", lambda r, l: r.sale_price_uzs),
    Col("cost_uzs", "Себестоимость (UZS)", "Tannarx (UZS)", "money", lambda r, l: r.purchase_price_uzs_snapshot),
    Col("profit", "Прибыль (UZS)", "Foyda (UZS)", "money", lambda r, l: r.profit_uzs),
    Col("status", "Статус", "Holat", "text", lambda r, l: _enum("sale_status", r.status, l)),
    Col("comment", "Комментарий", "Izoh", "text", lambda r, l: r.comment or ""),
    Col("created", "Создано", "Yaratilgan", "datetime", lambda r, l: r.created_at),
]

_DEVICE_COLS: list[Col] = [
    Col("brand", "Бренд", "Brend", "text", lambda r, l: r.brand),
    Col("model", "Модель", "Model", "text", lambda r, l: r.model),
    Col("imei", "IMEI", "IMEI", "text", lambda r, l: r.imei or ""),
    Col("serial", "Серийный №", "Seriya №", "text", lambda r, l: r.serial or ""),
    Col("category", "Категория", "Toifa", "text", lambda r, l: _enum("category", r.category, l)),
    Col("condition", "Состояние", "Holati", "text", lambda r, l: _enum("condition", r.condition, l)),
    Col("status", "Статус", "Holat", "text", lambda r, l: _enum("device_status", r.status, l)),
    Col("buy_price_uzs", "Цена закупа (UZS)", "Qabul narxi (UZS)", "money", lambda r, l: r.price_uzs),
    Col("buy_date", "Дата закупа", "Qabul sanasi", "date", lambda r, l: r.purchase_date),
    Col("notes", "Заметки", "Eslatmalar", "text", lambda r, l: r.notes or ""),
    Col("created", "Создано", "Yaratilgan", "datetime", lambda r, l: r.created_at),
]

_PURCHASE_COLS: list[Col] = [
    Col("date", "Дата", "Sana", "date", lambda r, l: r.purchase_date),
    Col("brand", "Бренд", "Brend", "text", lambda r, l: r.brand),
    Col("model", "Модель", "Model", "text", lambda r, l: r.model),
    Col("imei", "IMEI", "IMEI", "text", lambda r, l: r.imei or ""),
    Col("seller", "Продавец", "Sotuvchi", "text", lambda r, l: r.seller_name),
    Col("seller_phone", "Телефон", "Telefon", "text", lambda r, l: r.seller_phone or ""),
    Col("currency", "Валюта", "Valyuta", "text", lambda r, l: r.currency),
    Col("price_uzs", "Цена (UZS)", "Narx (UZS)", "money", lambda r, l: r.price_uzs),
    Col("price_usd", "Цена (USD)", "Narx (USD)", "money", lambda r, l: r.price_usd),
    Col("rate", "Курс", "Kurs", "money", lambda r, l: r.exchange_rate),
    Col("comment", "Комментарий", "Izoh", "text", lambda r, l: r.comment or ""),
    Col("created", "Создано", "Yaratilgan", "datetime", lambda r, l: r.created_at),
]

_REGISTRY: dict[str, list[Col]] = {
    "sales": _SALES_COLS,
    "devices": _DEVICE_COLS,
    "purchases": _PURCHASE_COLS,
}

ENTITIES = tuple(_REGISTRY.keys())


def columns_for(entity: str, *, lang: str = "ru") -> list[dict[str, str]]:
    """Metadata for the column-picker UI: ordered key/label/type per column."""
    return [
        {"key": c.key, "label": c.header(lang), "type": c.type}
        for c in _REGISTRY.get(entity, [])
    ]


# ── Row fetchers (shop-scoped) ──────────────────────────────────────────


async def fetch(
    db: AsyncSession,
    *,
    shop_id: int,
    entity: str,
    date_from: date | None,
    date_to: date | None,
) -> Sequence[Row]:
    """Fetch all rows for ``entity`` in the shop, optionally bounded by date.

    The select lists every column any registered Col might read, labelled so
    ``Col.get`` can pull by attribute name regardless of which subset the user
    exported.
    """
    if entity == "sales":
        stmt = (
            select(
                Sale.sale_date, Device.brand, Device.model, Device.imei,
                Device.category, Device.condition, Sale.buyer_name,
                Sale.buyer_phone, Sale.sale_type, Sale.sale_price_uzs,
                Sale.purchase_price_uzs_snapshot, Sale.profit_uzs, Sale.status,
                Sale.comment, Sale.created_at,
            )
            .join(Device, Device.id == Sale.device_id)
            .where(Sale.shop_id == shop_id)
            .order_by(Sale.sale_date.desc(), Sale.id.desc())
        )
        if date_from is not None:
            stmt = stmt.where(Sale.sale_date >= date_from)
        if date_to is not None:
            stmt = stmt.where(Sale.sale_date <= date_to)
    elif entity == "devices":
        stmt = (
            select(
                Device.brand, Device.model, Device.imei, Device.serial,
                Device.category, Device.condition, Device.status,
                Purchase.price_uzs, Purchase.purchase_date, Device.notes,
                Device.created_at,
            )
            .join(Purchase, Purchase.device_id == Device.id)
            .where(Device.shop_id == shop_id)
            .order_by(Device.created_at.desc(), Device.id.desc())
        )
        if date_from is not None:
            stmt = stmt.where(Purchase.purchase_date >= date_from)
        if date_to is not None:
            stmt = stmt.where(Purchase.purchase_date <= date_to)
    elif entity == "purchases":
        stmt = (
            select(
                Purchase.purchase_date, Device.brand, Device.model, Device.imei,
                Purchase.seller_name, Purchase.seller_phone, Purchase.currency,
                Purchase.price_uzs, Purchase.price_usd, Purchase.exchange_rate,
                Purchase.comment, Purchase.created_at,
            )
            .join(Device, Device.id == Purchase.device_id)
            .where(Purchase.shop_id == shop_id)
            .order_by(Purchase.purchase_date.desc(), Purchase.id.desc())
        )
        if date_from is not None:
            stmt = stmt.where(Purchase.purchase_date >= date_from)
        if date_to is not None:
            stmt = stmt.where(Purchase.purchase_date <= date_to)
    else:
        raise ValueError(f"unknown export entity: {entity!r}")

    return (await db.execute(stmt)).all()


# ── Workbook builder ────────────────────────────────────────────────────


def _cell_value(col: Col, row: Row, lang: str) -> Any:
    value = col.get(row, lang)
    if col.type == "money":
        return float(value) if value is not None else 0.0
    if col.type == "int":
        return int(value) if value is not None else 0
    # openpyxl can't write tz-aware datetimes — show them in Tashkent local
    # time with the tzinfo stripped (stored UTC-aware per DateTime(timezone=True)).
    if (
        col.type == "datetime"
        and isinstance(value, datetime)
        and value.tzinfo is not None
    ):
        return to_tashkent(value).replace(tzinfo=None)
    return value


def build_xlsx(
    entity: str, rows: Sequence[Row], selected: Sequence[str], lang: str = "ru"
) -> bytes:
    """Build a flat .xlsx of ``rows`` with only ``selected`` columns, in order.

    Unknown keys in ``selected`` are silently dropped; if nothing valid is
    left we fall back to every column for the entity so the user never gets an
    empty sheet from a typo.
    """
    by_key = {c.key: c for c in _REGISTRY.get(entity, [])}
    cols = [by_key[k] for k in selected if k in by_key]
    if not cols:
        cols = list(_REGISTRY.get(entity, []))

    wb = Workbook()
    ws = wb.active
    ws.title = entity

    bold = Font(bold=True)
    for ci, col in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value=col.header(lang)).font = bold

    for ri, row in enumerate(rows, start=2):
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci, value=_cell_value(col, row, lang))
            if col.type == "money":
                cell.number_format = _MONEY_FMT
            elif col.type == "date":
                cell.number_format = _DATE_FMT
            elif col.type == "datetime":
                cell.number_format = _DATETIME_FMT

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    for ci, col in enumerate(cols, start=1):
        width = 22 if col.type in ("money", "datetime") else 16
        ws.column_dimensions[get_column_letter(ci)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
